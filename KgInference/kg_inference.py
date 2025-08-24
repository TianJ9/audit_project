'''
    Part3: 业数图谱推理
    输入：
        判定逻辑；业数图谱
    输出：
        对于来源为SQL问数的数据：表中的关键字段
        对于来源为用户上载的数据：需要用户上载的数据

'''

import json
import re
from openai import OpenAI
from openpyxl import load_workbook
import os

client = OpenAI(
    base_url="https://api.deepseek.com",
    api_key="",
)

model = "deepseek-reasoner"
# model="deepseek-chat"
file_path = "KnowledgeGraph.xlsx"
output_path = f"fields_from_{model}.json"
LOGIC = "投产日期 + 60 > 竣工结算审定日期"


# 可以按需改写为本地部署的生成方式
def model_gen(prompt):
    result = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "你是一个专业的审计专家，擅长分析各种审计项目并进行异常排查。"},
            {"role": "user", "content": prompt}
        ],
        stream=False,
        temperature=0.0
    )
    response = result.choices[0].message.content
    # print(response)
    return response


def clean_model_json(text: str):
    trans_map = str.maketrans({
        '（': '(',
        '）': ')',
        '，': ',',
        '：': ':',
        '“': '"',
        '”': '"',
        '’': "'",
        '‘': "'"
    })
    cleaned = text.translate(trans_map)

    match = re.search(r'\[.*\]', cleaned, flags=re.S)
    if match:
        cleaned = match.group(0)

    cleaned = re.sub(r',\s*([}\]])', r'\1', cleaned)

    try:
        return cleaned
    except json.JSONDecodeError as e:
        print("JSON 解析失败:", e)
        print("清洗后的文本:", cleaned)
        return None


# Step1 根据判定逻辑选取流程
def get_process_from_kg():
    wb = load_workbook(file_path)
    sheet = wb.active

    processes = ""
    seen = set()

    for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
        cnt = 1
        col1 = row[0]
        col2 = row[1]

        if col1 is None:
            continue

        # 如果第二列有值，就拼接
        if col2:
            combined = f"{cnt}. {col1} — {col2}\n"
        else:
            combined = str(col1)

        if combined not in seen:
            processes += combined
            seen.add(combined)
            cnt += 1

    return processes


def choose_process(LOGIC):
    PROMPT_TEMPLATE = '''
    你的任务是：根据给定的判别逻辑，从给定的流程列表中选出**包含判别逻辑所需任意字段**的流程。

    步骤：
    1. 从判别逻辑中识别出所有涉及的字段名称。
    2. 在流程列表的“描述”中查找这些字段名称（或常见的同义字段）。
    3. 如果流程描述中包含判别逻辑所需的任意一个字段，就将该流程选出。
    4. 严格按给定的流程列表匹配，不要推测流程中有未列出的字段。
    5. 输出匹配的流程名称和描述。
    
    输入：
    - 判别逻辑："{logic}"
    - 流程列表：
      {processes}
    
    输出要求：
    - 使用JSON数组，每个元素包含"name"和"description"两个字段
    - 数组中按给定列表的顺序排列匹配的流程
    - 不要输出不在列表中的流程
    '''

    processes = get_process_from_kg()
    prompt = PROMPT_TEMPLATE.format(logic=LOGIC, processes=json.dumps(processes))
    while True:
        # response = model_gen(prompt).replace("'''", "").replace("json", "").strip()
        response = clean_model_json(model_gen(prompt))
        try:
            chosen_processes = json.loads(response)
            break
        except Exception:
            print("Unformatted response, try again")
    print(f"chosen processes: {chosen_processes}\n")
    return chosen_processes


# Step2 根据选定逻辑去对应的表中查字段；返回需要SQL查数的字段，以及需要用户上载的数据字段
def choose_field(book_name, logic):
    fields = {book_name: []}
    load_file = ""
    for file in os.listdir("datasets"):
        if file.lower() == f"odps.{book_name}.xlsx".lower():
            load_file = file
            print("Choosing fields from dataset:", file)

    wb = load_workbook(f"datasets/{load_file}")
    sheet = wb.active
    candidates = []

    TEMPLATE = '''
    你的任务是：根据给定的判定逻辑，从给定的字段列表中找出与逻辑相关的字段。

    步骤：
    1. 从判定逻辑中识别出所有涉及的字段名称（包括隐含的字段，例如时间、金额等）。
    2. 在字段列表中查找这些字段名称或它们的同义词、近义表达。
    3. 如果字段名称或描述与判定逻辑中涉及的字段相关，就将该字段选出。
    4. 严格根据给定字段列表匹配，不要推测存在未列出的字段。
    5. 输出匹配的字段名称及描述。
    
    输入：
    - 判定逻辑："{logic}"
    - 字段列表：
      {fields}
    
    输出要求：
    - 使用 JSON 数组，每个元素包含"field"和"description"两个字段
    - 数组顺序与给定列表一致
    - 不要输出不在列表中的字段
    - 只输出与判定逻辑相关的字段
    '''

    for row in sheet.iter_rows(min_row=1, values_only=True):
        cnt = 1
        col1 = row[0]
        col3 = row[2]  # 第三列索引为2

        combined = f"{cnt}. {col1 if col1 is not None else ''} — {col3 if col3 is not None else ''}\n"
        candidates.append(combined)
        cnt += 1

    prompt = TEMPLATE.format(logic=logic, fields=json.dumps(candidates))

    field_and_desc = json.loads(clean_model_json(model_gen(prompt)))
    print("field and desc:", field_and_desc)
    if field_and_desc:
        fields[book_name] = [item["field"] for item in field_and_desc]
    print(f"fields: {fields}")
    return fields


def infer_field(chosen_processes, logic):
    wb = load_workbook(file_path)
    sheet = wb.active
    output_fields = {"sql": [], "user": []}
    for item in chosen_processes:
        name = item["name"]
        description = item["description"]

        col1_to_col2 = {}
        for row in sheet.iter_rows(min_row=1, values_only=True):
            col1, col2 = row[0], row[1]
            if col1 and col2 and col1 not in col1_to_col2:
                col1_to_col2[col1] = col2

        for row in sheet.iter_rows(min_row=1, values_only=True):
            # print("row ", row)
            col1 = row[0]  # 第一列 流程名称
            col2 = row[1]  # 第二列 流程中的字段
            col4 = row[3]  # 第四列 数据获取方式(SQL/用户)

            # 处理描述为空的情况
            if not col2 and col1 in col1_to_col2:
                col2 = col1_to_col2[col1]

            # 找到对应的行
            # col1 == name or
            if col2.replace("（", "(").replace("）", ")") == description.replace("（", "(").replace("）", ")"):
                if "SQL" in col4:
                    book_name = row[4]
                    try:
                        chosen_fields = choose_field(book_name, logic)
                        output_fields["sql"].append(chosen_fields)
                    except Exception as e:
                        print("Data Sheet not available, pass")
                else:
                    if description not in output_fields["user"]:
                        output_fields["user"].append(description)
    return output_fields


def main():
    print("Step1: Choose Process")
    chosen_processes = choose_process(LOGIC)
    print("Step2: Infer Fields")
    fields = infer_field(chosen_processes, LOGIC)

    output = {"Chosen Processes": chosen_processes, "Fields": fields}
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=4)



if __name__ == "__main__":
    main()

