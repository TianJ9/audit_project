'''
    API Part3 问数
    输入：
        审计风险点；业数图谱
    输出：
        每个风险点的判断逻辑（比较哪些数据）
        取数的SQL脚本

'''

import json
import re
import openpyxl
from openai import OpenAI
from openpyxl import load_workbook
import pandas as pd
from jinja2 import Template
import random
import networkx as nx
import matplotlib.pyplot as plt


client = OpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key="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",
)

model = "deepseek/deepseek-chat-v3-0324"

# client = OpenAI(
#     base_url="https://api.deepseek.com",
#     api_key="",
# )
# model="deepseek-chat"


# 可以按需改写为本地部署的生成方式
def model_gen(prompt):
    result = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "你是一个专业的审计专家，擅长分析各种审计项目并进行异常排查。"},
            {"role": "user", "content": prompt}
        ],
        stream=False,
        temperature=0.0
    )
    response = result.choices[0].message.content
    # print(response)
    return response


def clean_model_json(text: str):
    trans_map = str.maketrans({
        '（': '(',
        '）': ')',
        '，': ',',
        '：': ':',
        '“': '"',
        '”': '"',
        '’': "'",
        '‘': "'"
    })
    cleaned = text.translate(trans_map)

    match = re.search(r'\[.*\]', cleaned, flags=re.S)
    if match:
        cleaned = match.group(0)

    cleaned = re.sub(r',\s*([}\]])', r'\1', cleaned)

    try:
        return cleaned
    except json.JSONDecodeError as e:
        print("JSON 解析失败:", e)
        print("清洗后的文本:", cleaned)
        return None


file_path = "KnowledgeGraph.xlsx"
output_path = f"fields_from_{model}.json"



print("API 3: 根据风险点找到需要比对的字段，同时生成取数SQL脚本")
print(f"使用{model}模型演示中")


# Step1 遍历1.1中的数据表，并
def find_risk_rows(file_path, search_value):
    print("取数Step1: 获取风险点对应描述、判定逻辑和判定参数")
    # 1. 读入文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 2. 查找第二行中的目标列索引
    target_headers = ["审计风险点", "审计风险描述", "政策制度及管理办法", "风险判定逻辑", "判定参数", "中台共享层表单编码"]

    col_indices = {}

    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col).value
        if cell_value in target_headers:
            col_indices[cell_value] = col

    print("找到的列索引：", col_indices)

    # 检查是否找齐
    if len(col_indices) < len(target_headers):
        print("警告：部分目标列没有找到！")

    # 3. 遍历对应列，查找匹配行
    matched_rows = []
    for header, col in col_indices.items():
        for row in range(3, ws.max_row + 1):  # 从第3行开始，因为第2行是表头
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == search_value:
                matched_rows.append(row)

    matched_data = []
    # 4. 输出结果
    for row in matched_rows:
        row_data = {header: ws.cell(row=row, column=col_indices[header]).value for header in col_indices}
        print(f"匹配行 {row} 的数据: {row_data}")
        matched_data.append(row_data)
    return matched_rows, matched_data


# Step2 针对每个风险点，输出判别逻辑：针对每个风险点，拿什么字段和什么字段对比
def choose_process(LOGIC):
    print("取数Step2: 根据风险点描述与逻辑，找出需要比对的字段")
    PROMPT_TEMPLATE = '''
    你的任务是：根据给定的判定逻辑，给定需要判别的字段，如
    
    - 示例判别逻辑：“资产转资金额/采购订单的设备购置价格>3倍”
    - 示例输出：["资产转资金额", "采购订单的设备购置价格"]
    
    输出要求：
    - 使用列表输出，每个元素为一个用于判别比较的字段
    - 严格基于判别逻辑，不要编造
    
    输入：
    - 判别逻辑："{logic}"
    '''

    prompt = PROMPT_TEMPLATE.format(logic=LOGIC)
    while True:
        # response = model_gen(prompt).replace("'''", "").replace("json", "").strip()
        response = clean_model_json(model_gen(prompt))
        try:
            chosen_fields = json.loads(response)
            break
        except Exception:
            print("Unformatted response, try again")
    print(f"chosen fields: {chosen_fields}\n")
    return chosen_fields


# Step3 建立图，返回图和推理路径
# NOTE 仅为演示进行的简化实现
def infer_graph():
    print("取数Step3: 读取数据并建立图谱")
    # TABLE_NAMES = ["PRPS", "PROJ", "BKPF", "BSEG", "CDPOS"]

    def remove_chinese(text: str) -> str:
        """去掉字符串中的所有中文字符"""
        if pd.isna(text):
            return ""
        return re.sub(r'[\u4e00-\u9fff]', '', str(text))

    df1 = pd.read_excel("data/1.2.xls", sheet_name="数据表字段信息清单", header=None)

    dict1 = {}
    for _, row in df1.iterrows():
        col4 = row[3]  # 第四列（index 从0开始，所以是3）
        col6 = row[5]  # 第六列（index=5）
        if pd.isna(col4) or pd.isna(col6):
            continue
        dict1.setdefault(col4, []).append(col6)

    # 2. 读取 2.xlsx 的 sheet2
    df2 = pd.read_excel("data/GraphResult.xlsx", sheet_name="数据清单", header=None)

    dict2 = {}
    # 遍历第2、3、4、6行（Excel行号，pandas index 从0开始，所以对应 index=1,2,3,5）
    for i in [1, 2, 3, 5]:
        row = df2.iloc[i]
        col3 = row[2]  # 第三列 (index=2)
        col4 = row[3]  # 第四列 (index=3)

        if pd.isna(col3) or pd.isna(col4):
            continue

        # 处理第三列：去掉 "表："
        key = str(col3).replace("表：", "").strip()

        # 处理第四列：按 "、" 分割并去掉中文
        values = [remove_chinese(x).strip() for x in str(col4).split("、") if x.strip()]
        values = [v for v in values if v]  # 去掉空字符串

        dict2[key] = values

    # 3. 对第一个字典 dict1 进行处理
    result_dict = {}
    for key, value1 in dict1.items():
        if key in dict2:
            value2 = dict2[key]
            # 先保留出现在 value2 中的元素
            filtered = [v for v in value1 if v in value2]
            # 加入前五个不在 value2 中的元素
            extra = [v for v in value1 if v not in value2][:5]
            result_dict[key] = filtered + extra
            result_dict[key] = list(set(result_dict[key]))
            random.shuffle(result_dict[key])
        else:
            # del result_dict[key]
            # result_dict[key] = value1  # 如果 dict2 中没有该 key，则保持原样
            continue
    print(f"result_dict: {result_dict}")


    # 创建图
    # G = nx.Graph()
    # # 添加节点和边
    # for table, fields in result_dict.items():
    #     G.add_node(table, type="table")  # 表节点
    #     for field in fields:
    #         G.add_node(field, type="field")  # 字段节点
    #         G.add_edge(table, field)  # 表 -> 字段
    #
    # # 布局
    # pos = nx.spring_layout(G, k=0.5, iterations=50)
    #
    # # 区分表节点和字段节点
    # table_nodes = [n for n, d in G.nodes(data=True) if d['type'] == 'table']
    # field_nodes = [n for n, d in G.nodes(data=True) if d['type'] == 'field']
    #
    # # 绘制
    # plt.figure(figsize=(12, 8))
    # nx.draw_networkx_nodes(G, pos, nodelist=table_nodes, node_color="lightblue", node_size=1000, label="Tables")
    # nx.draw_networkx_nodes(G, pos, nodelist=field_nodes, node_color="lightgreen", node_size=600, label="Fields")
    # nx.draw_networkx_edges(G, pos, alpha=0.5)
    # nx.draw_networkx_labels(G, pos, font_size=10)
    #
    # plt.legend()
    # plt.axis("off")
    # plt.show()
    return result_dict

# Step4 通过直接取数，完成表格对应数据的映射，获取用于生成SQL的字段
def find_fields_by_table(table_name, target_fields):
    print("取数Step4: 根据源数据进行图谱推理，找到生成SQL所需参数")
    """
    file_path: Excel文件路径
    table_name: 要匹配的表名（第8列）
    target_fields: 目标字段列表，用于匹配第9列
    """

    # file_path = "/ Users / pantianjun / Desktop / audit_project / Risk2SQL/data/SourceData.xlsx"
    file_path = "./data/SourceData.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    matched_rows = []

    # 1. 遍历第8列，找到表名匹配的行
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=8).value
        if cell_value == table_name:
            matched_rows.append(row)

    print(f"匹配表名 '{table_name}' 的行号: {matched_rows}")

    # 2. 遍历第9列，查找目标字段
    results = []
    for row in matched_rows:
        cell_value = ws.cell(row=row, column=9).value
        if cell_value is None:
            continue
        for field in target_fields:
            if field in str(cell_value):  # 包含匹配
                results.append({
                    "目标字段": field,
                    "整单元格内容": cell_value
                })

    # 3. 输出结果
    filtered = []
    for r in results:
        if r not in filtered:
            filtered.append(r)

    print(filtered)
    return filtered



# Step5 生成SQL
def generate_SQL(risk, target_fields, key_fields):
    print("取数Step5: 生成SQL文件")
    PROMPT_TEMPLATE = '''
    根据关键字段和目标字段，仿照下面的SQL脚本的格式，写一个取数脚本，只返回脚本
    
    关键字段：
    {key_fields}
    目标字段：
    {target_fields}
    脚本：
    SELECT  *
    FROM    (
                SELECT  proj.pspid AS prj_code    -- 项目编码
                        ,proj.post1 AS prj_name    -- 项目名称
                        ,prps.stufe AS prj_level    -- 层级
                        ,substr(prps.posid, 1, 14) single_prj_code_14    -- 前14位
                        ,prps.posid AS single_prj_code    -- 单体工程编码
                        ,prps.post1 AS single_prj_name    -- 单体工程名称
                        ,prps.usr08 AS operation_start_date    -- 转资凭证日期（竣工验收日期）
                        ,bkpf.bldat AS capitial_date    -- 投运转资日期
                        ,ROW_NUMBER() OVER(PARTITION BY proj.pspid ,proj.post1 ,prps.stufe ,substr(prps.posid, 1, 14) ,prps.posid ,prps.post1 ,prps.usr08 ORDER BY bkpf.bldat) row_index
                FROM    PRO_DWH_ERP_PRD.ods_erp_p00_sapsr3_PROJ proj
                LEFT JOIN pro_dwh_erp_prd.ods_erp_p00_sapsr3_prps PRPS
                ON      PROJ.PSPNR = prps.PSPHI  -- 外检关联(内码)
                AND     LENGTH(TRIM(PRPS.PSPNR)) > 0
                AND     prps.stufe IN (1, 2)
                AND     PRPS.mandt = '880'
                AND     PRPS.ds = MAX_PT('pro_dwh_erp_prd.ods_erp_p00_sapsr3_prps') LEFT
                JOIN    (    --过滤条件
                            SELECT  prj_code
                                    ,COUNT(DISTINCT single_prj_code_14) AS NUM    -- 差异数
                            FROM    (
                                        SELECT  DISTINCT proj.pspid AS prj_code    -- 项目编码
                                                ,substr(prps.posid, 1, 14) single_prj_code_14    -- 前14位
                                        FROM    PRO_DWH_ERP_PRD.ods_erp_p00_sapsr3_PROJ proj
                                        LEFT JOIN pro_dwh_erp_prd.ods_erp_p00_sapsr3_prps PRPS
                                        ON      PROJ.PSPNR = prps.PSPHI
                                        AND     LENGTH(TRIM(PRPS.PSPNR)) > 0
                                        AND     prps.stufe IN (1, 2)
                                        AND     PRPS.mandt = '880'
                                        AND     PRPS.ds = MAX_PT('pro_dwh_erp_prd.ods_erp_p00_sapsr3_prps')
                                        WHERE   proj.mandt = '880'
                                        AND     proj.ds = MAX_PT('PRO_DWH_ERP_PRD.ods_erp_p00_sapsr3_PROJ')
                                    ) 
                            GROUP BY prj_code
                        ) mid_tmp_filter_table
                ON      mid_tmp_filter_table.prj_code = proj.pspid
                LEFT JOIN (
                              SELECT  gjahr    -- 会计年度
                                      ,belnr    -- 会计凭证编码
                                      ,bukrs    -- 公司代码
                                      ,mandt    -- 客户端
                                      ,sgtxt    -- 外键
                                      ,hkont
                              FROM    pro_dwh_erp_prd.ods_erp_zltp_erp_bseg
                          ) bseg
                ON      substr(bseg.sgtxt, 5, 14) = substr(prps.posid, 1, 14)    -- 外键
                AND     bseg.mandt = '880'
                AND     substr(bseg.hkont, 1, 4) = '1601' 
                LEFT JOIN    pro_dwh_erp_prd.ods_erp_p00_sapsr3_bkpf bkpf
                ON      bseg.bukrs = bkpf.bukrs    -- 外键盘
                AND     bseg.belnr = bkpf.belnr
                AND     bseg.gjahr = bkpf.gjahr
                AND     bkpf.ds = max_pt('pro_dwh_erp_prd.ods_erp_p00_sapsr3_bkpf')
                WHERE   proj.mandt = '880'
                AND     proj.ds = MAX_PT('PRO_DWH_ERP_PRD.ods_erp_p00_sapsr3_PROJ')
                AND     CASE WHEN mid_tmp_filter_table.NUM == 1 THEN prps.stufe = 1 ELSE prps.stufe = 2 END
                AND     proj.pspid in 
                ('18138721004F', '18138721004B', '18138721000U')  
                -- '18138119006D'  
            ) 
    WHERE   row_index = 1
    ;
    '''

    prompt = PROMPT_TEMPLATE.format(key_fields=key_fields, target_fields=target_fields)
    while True:
        try:
            response = model_gen(prompt)
            break
        except Exception as e:
            print(e)
    # output_name = "_".join(target_fields)
    # output_path = f"{risk}_{model}_{target_fields}.sql"
    # with open(output_path, "w", encoding="utf-8") as f:
    #     f.write(response)
    return response


def pipeline(risk):
    print(f"正在为风险点{risk}进行查数操作")
    file_path = "./data/KG.xlsm"  # 你的xlsx路径
    rows, rows_data = find_risk_rows(file_path, risk)
    output = []
    for cnt, item in enumerate(rows_data):
        logic = item["风险判定逻辑"]
        chosen_fields = choose_process(logic)

        epoch = {
            "风险点": risk,
            "审计风险描述": item["审计风险描述"],
            "政策制度及管理办法": item["政策制度及管理办法"],
            "风险判定逻辑": logic,
            "比对字段": chosen_fields,

        }
        output.append(epoch)
        print(item.keys())
        if "." in item["中台共享层表单编码"] and len(item["中台共享层表单编码"].splitlines()) > 1:
            table_names = [line.split('.', 1)[1] for line in item["中台共享层表单编码"].splitlines()]
        else:
            table_names = [item["中台共享层表单编码"]]

        for table_name in table_names:
            target_fields_for_sql = find_fields_by_table(table_name, chosen_fields)

            for instance in target_fields_for_sql:
                target_fields = instance["目标字段"]
                key_fields = instance["整单元格内容"]
                SQL = generate_SQL(risk, target_fields, key_fields)
                output.append(SQL)

    return output


def run_pipelines(risks):
    output = []
    for risk in risks:
        output.extend(pipeline(risk))

    return output


def main():
    # 输入风险点
    risks = [
        "预转资异常",
        "审定结算不及时"
    ]
    output = []
    graph_dict = infer_graph()

    for risk in risks:
        output.extend(pipeline(risk))
    with open("graph.json", "w", encoding="utf-8") as f:
        json.dump(graph_dict, f, ensure_ascii=False, indent=4)
    with open(f"{model}.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=4)

if __name__ == "__main__":
    main()
    # infer_graph()
